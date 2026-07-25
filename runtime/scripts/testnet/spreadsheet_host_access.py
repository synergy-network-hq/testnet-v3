#!/usr/bin/env python3
"""Run Synergy Testnet host commands using the operator credential workbook.

The workbook is the source of truth for host, SSH user, custom port, and
credential fields. This helper intentionally never prints credential values.
"""

from __future__ import annotations

import argparse
import os
import shlex
import subprocess
import sys
import tempfile
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_WORKBOOK = Path("/Users/devpup/Desktop/node-machine-credentials.xlsx")


@dataclass(frozen=True)
class HostRow:
    row_number: int
    node: str
    ssh_command: str
    ssh_user: str
    public_ip: str
    qrpc_port: str
    ws_port: str
    metrics_port: str
    password: str
    passphrase: str


def load_hosts(workbook: Path) -> dict[str, HostRow]:
    wb = load_workbook(workbook, data_only=True, read_only=True)
    rows = []
    for candidate in wb.worksheets:
        candidate_rows = list(candidate.iter_rows(values_only=True))
        if not candidate_rows:
            continue
        candidate_headers = [
            str(value).strip() if value is not None else ""
            for value in candidate_rows[0]
        ]
        if "Node" in candidate_headers and "Access Via SSH with" in candidate_headers:
            rows = candidate_rows
            break
    if not rows:
        rows = list(wb.active.iter_rows(values_only=True))
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    hosts: dict[str, HostRow] = {}

    def cell(row: tuple[object, ...], header: str, occurrence: int = 1) -> str:
        seen = 0
        for index, candidate in enumerate(headers):
            if candidate != header:
                continue
            seen += 1
            if seen == occurrence:
                value = row[index] if index < len(row) else None
                return str(value or "").strip()
        return ""

    for row_number, row in enumerate(rows[1:], start=2):
        node = cell(row, "Node")
        if not node:
            continue
        host = HostRow(
            row_number=row_number,
            node=node,
            ssh_command=cell(row, "Access Via SSH with", occurrence=1),
            ssh_user=cell(row, "SSH User") or cell(row, "User"),
            public_ip=cell(row, "Public IP"),
            qrpc_port=cell(row, "qRPC"),
            ws_port=cell(row, "WS"),
            metrics_port=cell(row, "Metrics"),
            password=cell(row, "User Password"),
            passphrase=cell(row, "SSH Passphrase"),
        )
        hosts[node.lower()] = host
        hosts[node.replace(" ", "").replace("-", "").lower()] = host
    return hosts


def sanitized_host_line(host: HostRow) -> str:
    return (
        f"spreadsheet_row_used=true row={host.row_number} node={host.node} "
        f"ssh={host.ssh_command!r} user={host.ssh_user!r} "
        f"public_ip={host.public_ip!r} qrpc={host.qrpc_port!r} "
        f"ws={host.ws_port!r} metrics={host.metrics_port!r}"
    )


def run_remote(
    host: HostRow,
    remote_command: str,
    timeout: int,
    password_auth: bool = False,
    remote_sudo_from_workbook: bool = False,
    extra_env: list[str] | None = None,
) -> int:
    if not host.ssh_command:
        print(f"missing SSH command for {host.node}", file=sys.stderr)
        return 2
    if password_auth:
        print(
            "refusing forced password-auth; workbook SSH command must be used exactly",
            file=sys.stderr,
        )
        return 2
    password = host.password or host.passphrase
    env = os.environ.copy()

    # Preserve the exact workbook SSH command and append only remote environment
    # values sourced from the same row plus the shell command.
    ssh_parts = shlex.split(host.ssh_command)
    env_items = [
        f"SYNERGY_SPREADSHEET_ROW={shlex.quote(str(host.row_number))}",
        f"SYNERGY_NODE={shlex.quote(host.node)}",
        f"SYNERGY_QRPC_PORT={shlex.quote(host.qrpc_port)}",
        f"SYNERGY_WS_PORT={shlex.quote(host.ws_port)}",
        f"SYNERGY_METRICS_PORT={shlex.quote(host.metrics_port)}",
    ]
    for item in extra_env or []:
        if "=" not in item:
            raise ValueError(f"remote env must be NAME=VALUE, got {item!r}")
        name, value = item.split("=", 1)
        if not name.replace("_", "").isalnum():
            raise ValueError(f"invalid remote env name {name!r}")
        env_items.append(f"{name}={shlex.quote(value)}")
    sudo_password = host.password or host.passphrase
    if remote_sudo_from_workbook and sudo_password:
        env_items.append(f"SYNERGY_REMOTE_SUDO_PASSWORD={shlex.quote(sudo_password)}")
    remote_env = " ".join(env_items)
    remote_shell = f"env {remote_env} bash -lc {shlex.quote(remote_command)}"
    print(sanitized_host_line(host), flush=True)
    askpass_path = None
    try:
        if password:
            fd, askpass_path = tempfile.mkstemp(prefix="synergy-ssh-askpass-", text=True)
            with os.fdopen(fd, "w", encoding="utf-8") as handle:
                handle.write("#!/bin/sh\n")
                handle.write('printf "%s\\n" "$SYNERGY_SSH_SECRET"\n')
            os.chmod(askpass_path, 0o700)
            env["SYNERGY_SSH_SECRET"] = password
            env["SSH_ASKPASS"] = askpass_path
            env["SSH_ASKPASS_REQUIRE"] = "force"
            env.setdefault("DISPLAY", ":0")
        try:
            completed = subprocess.run(
                [*ssh_parts, remote_shell],
                env=env,
                text=True,
                timeout=timeout,
                stdin=subprocess.DEVNULL,
            )
        except subprocess.TimeoutExpired:
            print(
                f"run timed out for node={host.node} after {timeout}s",
                file=sys.stderr,
            )
            return 124
        return completed.returncode
    finally:
        env.pop("SYNERGY_SSH_SECRET", None)
        if askpass_path:
            try:
                os.unlink(askpass_path)
            except FileNotFoundError:
                pass


def scp_command_parts(host: HostRow, password_auth: bool = False) -> list[str]:
    ssh_parts = shlex.split(host.ssh_command)
    if not ssh_parts or ssh_parts[0] != "ssh":
        raise ValueError(f"unsupported SSH command for scp conversion: {host.ssh_command}")
    scp_parts = ["scp"]
    if password_auth:
        if host.ssh_command.strip() == "ssh synergyvps":
            raise ValueError("refusing password-auth SSH options for raw synergyvps alias")
        scp_parts.extend(
            [
                "-o",
                "PreferredAuthentications=password",
                "-o",
                "PubkeyAuthentication=no",
                "-o",
                "IdentitiesOnly=yes",
            ]
        )
    index = 1
    remote_target = None
    while index < len(ssh_parts):
        part = ssh_parts[index]
        if part == "-p" and index + 1 < len(ssh_parts):
            scp_parts.extend(["-P", ssh_parts[index + 1]])
            index += 2
            continue
        if part.startswith("-"):
            scp_parts.append(part)
            index += 1
            continue
        remote_target = part
        index += 1
    if remote_target is None:
        raise ValueError(f"missing remote target in SSH command: {host.ssh_command}")
    return scp_parts, remote_target


def transfer_file(
    host: HostRow,
    direction: str,
    local_path: Path,
    remote_path: str,
    timeout: int,
    password_auth: bool = False,
) -> int:
    if password_auth:
        print(
            "refusing forced password-auth; workbook SSH command must be used exactly",
            file=sys.stderr,
        )
        return 2
    password = host.password or host.passphrase
    env = os.environ.copy()
    ssh_parts = shlex.split(host.ssh_command)
    askpass_path = None
    stdin_handle = None
    stdout_handle = None
    print(sanitized_host_line(host), flush=True)
    try:
        if password:
            fd, askpass_path = tempfile.mkstemp(prefix="synergy-ssh-askpass-", text=True)
            with os.fdopen(fd, "w", encoding="utf-8") as handle:
                handle.write("#!/bin/sh\n")
                handle.write('printf "%s\\n" "$SYNERGY_SSH_SECRET"\n')
            os.chmod(askpass_path, 0o700)
            env["SYNERGY_SSH_SECRET"] = password
            env["SSH_ASKPASS"] = askpass_path
            env["SSH_ASKPASS_REQUIRE"] = "force"
            env.setdefault("DISPLAY", ":0")
        if direction == "upload":
            stdin_handle = local_path.open("rb")
            remote_shell = f"cat > {shlex.quote(remote_path)}"
            stdout_target = None
        else:
            local_path.parent.mkdir(parents=True, exist_ok=True)
            stdout_handle = local_path.open("wb")
            remote_shell = f"cat {shlex.quote(remote_path)}"
            stdout_target = stdout_handle
        try:
            completed = subprocess.run(
                [*ssh_parts, remote_shell],
                env=env,
                text=False,
                timeout=timeout,
                stdin=stdin_handle or subprocess.DEVNULL,
                stdout=stdout_target,
            )
        except subprocess.TimeoutExpired:
            print(
                f"{direction} timed out for node={host.node} after {timeout}s",
                file=sys.stderr,
            )
            return 124
        return completed.returncode
    finally:
        if stdin_handle:
            stdin_handle.close()
        if stdout_handle:
            stdout_handle.close()
        env.pop("SYNERGY_SSH_SECRET", None)
        if askpass_path:
            try:
                os.unlink(askpass_path)
            except FileNotFoundError:
                pass


def stream_file_to_remote_command(
    host: HostRow,
    local_path: Path,
    remote_command: str,
    timeout: int,
    password_auth: bool = False,
    remote_sudo_from_workbook: bool = False,
    extra_env: list[str] | None = None,
) -> int:
    if not host.ssh_command:
        print(f"missing SSH command for {host.node}", file=sys.stderr)
        return 2
    if password_auth:
        print(
            "refusing forced password-auth; workbook SSH command must be used exactly",
            file=sys.stderr,
        )
        return 2
    password = host.password or host.passphrase
    env = os.environ.copy()
    ssh_parts = shlex.split(host.ssh_command)
    env_items = [
        f"SYNERGY_SPREADSHEET_ROW={shlex.quote(str(host.row_number))}",
        f"SYNERGY_NODE={shlex.quote(host.node)}",
        f"SYNERGY_QRPC_PORT={shlex.quote(host.qrpc_port)}",
        f"SYNERGY_WS_PORT={shlex.quote(host.ws_port)}",
        f"SYNERGY_METRICS_PORT={shlex.quote(host.metrics_port)}",
    ]
    for item in extra_env or []:
        if "=" not in item:
            raise ValueError(f"remote env must be NAME=VALUE, got {item!r}")
        name, value = item.split("=", 1)
        if not name.replace("_", "").isalnum():
            raise ValueError(f"invalid remote env name {name!r}")
        env_items.append(f"{name}={shlex.quote(value)}")
    sudo_password = host.password or host.passphrase
    if remote_sudo_from_workbook and sudo_password:
        env_items.append(f"SYNERGY_REMOTE_SUDO_PASSWORD={shlex.quote(sudo_password)}")
    remote_env = " ".join(env_items)
    remote_shell = f"env {remote_env} bash -lc {shlex.quote(remote_command)}"
    askpass_path = None
    stdin_handle = None
    print(sanitized_host_line(host), flush=True)
    try:
        if password:
            fd, askpass_path = tempfile.mkstemp(prefix="synergy-ssh-askpass-", text=True)
            with os.fdopen(fd, "w", encoding="utf-8") as handle:
                handle.write("#!/bin/sh\n")
                handle.write('printf "%s\\n" "$SYNERGY_SSH_SECRET"\n')
            os.chmod(askpass_path, 0o700)
            env["SYNERGY_SSH_SECRET"] = password
            env["SSH_ASKPASS"] = askpass_path
            env["SSH_ASKPASS_REQUIRE"] = "force"
            env.setdefault("DISPLAY", ":0")
        stdin_handle = local_path.open("rb")
        try:
            completed = subprocess.run(
                [*ssh_parts, remote_shell],
                env=env,
                text=False,
                timeout=timeout,
                stdin=stdin_handle,
            )
        except subprocess.TimeoutExpired:
            print(
                f"stream-run timed out for node={host.node} after {timeout}s",
                file=sys.stderr,
            )
            return 124
        return completed.returncode
    finally:
        if stdin_handle:
            stdin_handle.close()
        env.pop("SYNERGY_SSH_SECRET", None)
        if askpass_path:
            try:
                os.unlink(askpass_path)
            except FileNotFoundError:
                pass


def download_remote_command(
    host: HostRow,
    remote_command: str,
    local_path: Path,
    timeout: int,
    password_auth: bool = False,
    extra_env: list[str] | None = None,
) -> int:
    if not host.ssh_command:
        print(f"missing SSH command for {host.node}", file=sys.stderr)
        return 2
    if password_auth:
        print(
            "refusing forced password-auth; workbook SSH command must be used exactly",
            file=sys.stderr,
        )
        return 2
    password = host.password or host.passphrase
    env = os.environ.copy()
    ssh_parts = shlex.split(host.ssh_command)
    env_items = [
        f"SYNERGY_SPREADSHEET_ROW={shlex.quote(str(host.row_number))}",
        f"SYNERGY_NODE={shlex.quote(host.node)}",
        f"SYNERGY_QRPC_PORT={shlex.quote(host.qrpc_port)}",
        f"SYNERGY_WS_PORT={shlex.quote(host.ws_port)}",
        f"SYNERGY_METRICS_PORT={shlex.quote(host.metrics_port)}",
    ]
    for item in extra_env or []:
        if "=" not in item:
            raise ValueError(f"remote env must be NAME=VALUE, got {item!r}")
        name, value = item.split("=", 1)
        if not name.replace("_", "").isalnum():
            raise ValueError(f"invalid remote env name {name!r}")
        env_items.append(f"{name}={shlex.quote(value)}")
    remote_env = " ".join(env_items)
    remote_shell = f"env {remote_env} bash -lc {shlex.quote(remote_command)}"
    askpass_path = None
    stdout_handle = None
    print(sanitized_host_line(host), file=sys.stderr, flush=True)
    try:
        if password:
            fd, askpass_path = tempfile.mkstemp(prefix="synergy-ssh-askpass-", text=True)
            with os.fdopen(fd, "w", encoding="utf-8") as handle:
                handle.write("#!/bin/sh\n")
                handle.write('printf "%s\\n" "$SYNERGY_SSH_SECRET"\n')
            os.chmod(askpass_path, 0o700)
            env["SYNERGY_SSH_SECRET"] = password
            env["SSH_ASKPASS"] = askpass_path
            env["SSH_ASKPASS_REQUIRE"] = "force"
            env.setdefault("DISPLAY", ":0")
        local_path.parent.mkdir(parents=True, exist_ok=True)
        stdout_handle = local_path.open("wb")
        try:
            completed = subprocess.run(
                [*ssh_parts, remote_shell],
                env=env,
                text=False,
                timeout=timeout,
                stdin=subprocess.DEVNULL,
                stdout=stdout_handle,
            )
        except subprocess.TimeoutExpired:
            print(
                f"download-command timed out for node={host.node} after {timeout}s",
                file=sys.stderr,
            )
            return 124
        return completed.returncode
    finally:
        if stdout_handle:
            stdout_handle.close()
        env.pop("SYNERGY_SSH_SECRET", None)
        if askpass_path:
            try:
                os.unlink(askpass_path)
            except FileNotFoundError:
                pass


def pipe_remote_to_remote(
    source: HostRow,
    source_command: str,
    target: HostRow,
    target_command: str,
    timeout: int,
    source_extra_env: list[str] | None = None,
    target_extra_env: list[str] | None = None,
) -> int:
    def prepare(host: HostRow, remote_command: str, extra_env: list[str] | None):
        password = host.password or host.passphrase
        env = os.environ.copy()
        ssh_parts = shlex.split(host.ssh_command)
        env_items = [
            f"SYNERGY_SPREADSHEET_ROW={shlex.quote(str(host.row_number))}",
            f"SYNERGY_NODE={shlex.quote(host.node)}",
            f"SYNERGY_QRPC_PORT={shlex.quote(host.qrpc_port)}",
            f"SYNERGY_WS_PORT={shlex.quote(host.ws_port)}",
            f"SYNERGY_METRICS_PORT={shlex.quote(host.metrics_port)}",
        ]
        for item in extra_env or []:
            if "=" not in item:
                raise ValueError(f"remote env must be NAME=VALUE, got {item!r}")
            name, value = item.split("=", 1)
            if not name.replace("_", "").isalnum():
                raise ValueError(f"invalid remote env name {name!r}")
            env_items.append(f"{name}={shlex.quote(value)}")
        remote_env = " ".join(env_items)
        remote_shell = f"env {remote_env} bash -lc {shlex.quote(remote_command)}"
        askpass_path = None
        if password:
            fd, askpass_path = tempfile.mkstemp(prefix="synergy-ssh-askpass-", text=True)
            with os.fdopen(fd, "w", encoding="utf-8") as handle:
                handle.write("#!/bin/sh\n")
                handle.write('printf "%s\\n" "$SYNERGY_SSH_SECRET"\n')
            os.chmod(askpass_path, 0o700)
            env["SYNERGY_SSH_SECRET"] = password
            env["SSH_ASKPASS"] = askpass_path
            env["SSH_ASKPASS_REQUIRE"] = "force"
            env.setdefault("DISPLAY", ":0")
        return [*ssh_parts, remote_shell], env, askpass_path

    source_proc = None
    target_proc = None
    askpass_paths: list[str] = []
    print("source_" + sanitized_host_line(source), file=sys.stderr, flush=True)
    print("target_" + sanitized_host_line(target), file=sys.stderr, flush=True)
    try:
        source_args, source_env, source_askpass = prepare(
            source, source_command, source_extra_env
        )
        target_args, target_env, target_askpass = prepare(
            target, target_command, target_extra_env
        )
        askpass_paths.extend(path for path in (source_askpass, target_askpass) if path)
        source_proc = subprocess.Popen(
            source_args,
            env=source_env,
            stdin=subprocess.DEVNULL,
            stdout=subprocess.PIPE,
        )
        assert source_proc.stdout is not None
        target_proc = subprocess.Popen(
            target_args,
            env=target_env,
            stdin=source_proc.stdout,
        )
        source_proc.stdout.close()
        try:
            target_rc = target_proc.wait(timeout=timeout)
            source_rc = source_proc.wait(timeout=30)
        except subprocess.TimeoutExpired:
            print(
                f"pipe-run timed out source={source.node} target={target.node} after {timeout}s",
                file=sys.stderr,
            )
            for proc in (target_proc, source_proc):
                if proc and proc.poll() is None:
                    proc.kill()
            return 124
        if source_rc != 0:
            print(
                f"pipe-run source failed node={source.node} exit={source_rc}",
                file=sys.stderr,
            )
            return source_rc
        if target_rc != 0:
            print(
                f"pipe-run target failed node={target.node} exit={target_rc}",
                file=sys.stderr,
            )
            return target_rc
        return 0
    finally:
        for proc in (target_proc, source_proc):
            if proc and proc.poll() is None:
                proc.kill()
        for path in askpass_paths:
            try:
                os.unlink(path)
            except FileNotFoundError:
                pass


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    subparsers = parser.add_subparsers(dest="command", required=True)

    inventory = subparsers.add_parser("inventory")
    inventory.add_argument("--nodes", nargs="*")

    run = subparsers.add_parser("run")
    run.add_argument("node")
    run.add_argument("remote_command")
    run.add_argument("--timeout", type=int, default=60)
    run.add_argument("--remote-env", action="append", default=[])
    run.add_argument("--remote-sudo-from-workbook", action="store_true")
    run.add_argument(
        "--password-auth",
        action="store_true",
        help="Suppress local SSH keys for non-synergyvps hosts when exact SSH fails from key flood.",
    )

    run_file = subparsers.add_parser("run-file")
    run_file.add_argument("node")
    run_file.add_argument("script_path", type=Path)
    run_file.add_argument("--timeout", type=int, default=60)
    run_file.add_argument("--remote-env", action="append", default=[])
    run_file.add_argument("--remote-sudo-from-workbook", action="store_true")
    run_file.add_argument(
        "--password-auth",
        action="store_true",
        help="Suppress local SSH keys for non-synergyvps hosts when exact SSH fails from key flood.",
    )

    download = subparsers.add_parser("download")
    download.add_argument("node")
    download.add_argument("remote_path")
    download.add_argument("local_path", type=Path)
    download.add_argument("--timeout", type=int, default=120)
    download.add_argument("--password-auth", action="store_true")

    upload = subparsers.add_parser("upload")
    upload.add_argument("node")
    upload.add_argument("local_path", type=Path)
    upload.add_argument("remote_path")
    upload.add_argument("--timeout", type=int, default=120)
    upload.add_argument("--password-auth", action="store_true")

    stream_run = subparsers.add_parser("stream-run")
    stream_run.add_argument("node")
    stream_run.add_argument("local_path", type=Path)
    stream_run.add_argument("remote_command")
    stream_run.add_argument("--timeout", type=int, default=120)
    stream_run.add_argument("--remote-env", action="append", default=[])
    stream_run.add_argument("--remote-sudo-from-workbook", action="store_true")
    stream_run.add_argument("--password-auth", action="store_true")

    download_command = subparsers.add_parser("download-command")
    download_command.add_argument("node")
    download_command.add_argument("remote_command")
    download_command.add_argument("local_path", type=Path)
    download_command.add_argument("--timeout", type=int, default=120)
    download_command.add_argument("--remote-env", action="append", default=[])
    download_command.add_argument("--password-auth", action="store_true")

    pipe_run = subparsers.add_parser("pipe-run")
    pipe_run.add_argument("source_node")
    pipe_run.add_argument("source_command")
    pipe_run.add_argument("target_node")
    pipe_run.add_argument("target_command")
    pipe_run.add_argument("--timeout", type=int, default=120)
    pipe_run.add_argument("--source-remote-env", action="append", default=[])
    pipe_run.add_argument("--target-remote-env", action="append", default=[])

    args = parser.parse_args()
    hosts = load_hosts(args.workbook)

    if args.command == "inventory":
        selected = args.nodes or sorted({host.node for host in hosts.values()})
        for node in selected:
            host = hosts.get(node.lower()) or hosts.get(
                node.replace(" ", "").replace("-", "").lower()
            )
            if host is None:
                print(f"missing workbook row for node={node}", file=sys.stderr)
                return 2
            print(sanitized_host_line(host))
        return 0

    if args.command == "run":
        host = hosts.get(args.node.lower()) or hosts.get(
            args.node.replace(" ", "").replace("-", "").lower()
        )
        if host is None:
            print(f"missing workbook row for node={args.node}", file=sys.stderr)
            return 2
        return run_remote(
            host,
            args.remote_command,
            args.timeout,
            args.password_auth,
            args.remote_sudo_from_workbook,
            args.remote_env,
        )

    if args.command == "run-file":
        host = hosts.get(args.node.lower()) or hosts.get(
            args.node.replace(" ", "").replace("-", "").lower()
        )
        if host is None:
            print(f"missing workbook row for node={args.node}", file=sys.stderr)
            return 2
        return run_remote(
            host,
            args.script_path.read_text(),
            args.timeout,
            args.password_auth,
            args.remote_sudo_from_workbook,
            args.remote_env,
        )

    if args.command in {"download", "upload"}:
        host = hosts.get(args.node.lower()) or hosts.get(
            args.node.replace(" ", "").replace("-", "").lower()
        )
        if host is None:
            print(f"missing workbook row for node={args.node}", file=sys.stderr)
            return 2
        return transfer_file(
            host,
            args.command,
            args.local_path,
            args.remote_path,
            args.timeout,
            args.password_auth,
        )

    if args.command == "stream-run":
        host = hosts.get(args.node.lower()) or hosts.get(
            args.node.replace(" ", "").replace("-", "").lower()
        )
        if host is None:
            print(f"missing workbook row for node={args.node}", file=sys.stderr)
            return 2
        return stream_file_to_remote_command(
            host,
            args.local_path,
            args.remote_command,
            args.timeout,
            args.password_auth,
            args.remote_sudo_from_workbook,
            args.remote_env,
        )

    if args.command == "download-command":
        host = hosts.get(args.node.lower()) or hosts.get(
            args.node.replace(" ", "").replace("-", "").lower()
        )
        if host is None:
            print(f"missing workbook row for node={args.node}", file=sys.stderr)
            return 2
        return download_remote_command(
            host,
            args.remote_command,
            args.local_path,
            args.timeout,
            args.password_auth,
            args.remote_env,
        )

    if args.command == "pipe-run":
        source = hosts.get(args.source_node.lower()) or hosts.get(
            args.source_node.replace(" ", "").replace("-", "").lower()
        )
        target = hosts.get(args.target_node.lower()) or hosts.get(
            args.target_node.replace(" ", "").replace("-", "").lower()
        )
        if source is None:
            print(f"missing workbook row for node={args.source_node}", file=sys.stderr)
            return 2
        if target is None:
            print(f"missing workbook row for node={args.target_node}", file=sys.stderr)
            return 2
        return pipe_remote_to_remote(
            source,
            args.source_command,
            target,
            args.target_command,
            args.timeout,
            args.source_remote_env,
            args.target_remote_env,
        )

    return 2


if __name__ == "__main__":
    raise SystemExit(main())
